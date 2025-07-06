"""Service for exporting data to Excel."""

import os
from datetime import date
from typing import Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

from app.services.index_service import index_service
from app.db.database import db
from app.config import get_settings

logger = logging.getLogger(__name__)
settings = get_settings()


class ExportService:
    """Service for exporting index data to Excel."""
    
    def __init__(self):
        self.export_dir = settings.export_dir
        self._ensure_export_directory()
    
    def _ensure_export_directory(self):
        """Ensure export directory exists."""
        if not os.path.exists(self.export_dir):
            os.makedirs(self.export_dir, exist_ok=True)
    
    def export_to_excel(self, start_date: date, end_date: date) -> Dict:
        """Export index data to Excel file."""
        logger.info(f"Exporting data from {start_date} to {end_date}")
        
        # Generate filename
        filename = f"index_data_{start_date}_{end_date}.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Add sheets with data
        self._add_performance_sheet(wb, start_date, end_date)
        self._add_compositions_sheet(wb, start_date, end_date)
        self._add_changes_sheet(wb, start_date, end_date)
        self._add_summary_sheet(wb, start_date, end_date)
        
        # Save workbook
        wb.save(filepath)
        
        # Get file size
        file_size_mb = os.path.getsize(filepath) / (1024 * 1024)
        
        return {
            "success": True,
            "file_path": filepath,
            "file_size_mb": round(file_size_mb, 2)
        }
    
    def _add_performance_sheet(self, wb: Workbook, start_date: date, end_date: date):
        """Add performance data sheet."""
        ws = wb.create_sheet("Performance")
        
        # Get performance data
        try:
            perf_data = index_service.get_performance(start_date, end_date)
        except ValueError as e:
            ws.append(["Error: " + str(e)])
            return
        
        # Headers
        headers = ["Date", "Index Value", "Daily Return (%)", "Cumulative Return (%)"]
        ws.append(headers)
        
        # Apply header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row in perf_data['performance_data']:
            ws.append([
                row['date'],
                row['value'],
                row['daily_return'],
                row['cumulative_return']
            ])
        
        # Format columns
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        
        # Add summary section
        ws.append([])  # Empty row
        ws.append(["Summary Statistics"])
        ws['A' + str(ws.max_row)].font = Font(bold=True)
        
        summary = perf_data['summary']
        ws.append(["Total Return (%)", summary['total_return']])
        ws.append(["Average Daily Return (%)", summary['average_daily_return']])
        ws.append(["Volatility (Daily %)", summary['volatility']])
        ws.append(["Max Daily Return (%)", summary['max_daily_return']])
        ws.append(["Min Daily Return (%)", summary['min_daily_return']])
        ws.append(["Sharpe Ratio (Annualized)", summary['sharpe_ratio']])
        
        # Format numbers
        for row in ws.iter_rows(min_row=2):
            if row[0].value and isinstance(row[0].value, date):
                row[0].number_format = 'YYYY-MM-DD'
            for cell in row[1:]:
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
    
    def _add_compositions_sheet(self, wb: Workbook, start_date: date, end_date: date):
        """Add daily compositions sheet."""
        ws = wb.create_sheet("Daily Compositions")
        
        # Get all trading dates
        with db.get_connection() as conn:
            dates_df = conn.execute("""
                SELECT DISTINCT date 
                FROM index_compositions 
                WHERE date >= ? AND date <= ?
                ORDER BY date
            """, (start_date, end_date)).pl()
        
        if dates_df.is_empty():
            ws.append(["No composition data available"])
            return
        
        # Headers
        headers = ["Date", "Ticker", "Name", "Weight (%)", "Market Cap (B)", "Sector", "Industry"]
        ws.append(headers)
        
        # Apply header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data for each date
        for comp_date in dates_df['date'].to_list():
            try:
                comp_data = index_service.get_composition(comp_date)
                for stock in comp_data['compositions']:
                    ws.append([
                        comp_date,
                        stock['ticker'],
                        stock['name'],
                        stock['weight'] * 100,
                        stock['market_cap'] / 1e9,  # Convert to billions
                        stock['sector'],
                        stock['industry']
                    ])
            except ValueError:
                continue
        
        # Format columns
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 25
        
        # Format numbers
        for row in ws.iter_rows(min_row=2):
            if row[0].value and isinstance(row[0].value, date):
                row[0].number_format = 'YYYY-MM-DD'
            if row[3].value and isinstance(row[3].value, (int, float)):
                row[3].number_format = '0.00'
            if row[4].value and isinstance(row[4].value, (int, float)):
                row[4].number_format = '#,##0.00'
    
    def _add_changes_sheet(self, wb: Workbook, start_date: date, end_date: date):
        """Add composition changes sheet."""
        ws = wb.create_sheet("Composition Changes")
        
        # Get changes data
        try:
            changes_data = index_service.get_composition_changes(start_date, end_date)
        except ValueError as e:
            ws.append(["Error: " + str(e)])
            return
        
        # Headers
        headers = ["Date", "Action", "Ticker", "Name", "Market Cap (B)"]
        ws.append(headers)
        
        # Apply header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for change in changes_data['changes']:
            change_date = change['date']
            
            # Add stocks added
            for stock in change['stocks_added']:
                ws.append([
                    change_date,
                    "Added",
                    stock['ticker'],
                    stock['name'],
                    stock['market_cap'] / 1e9
                ])
                # Color added rows green
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            # Add stocks removed
            for stock in change['stocks_removed']:
                ws.append([
                    change_date,
                    "Removed",
                    stock['ticker'],
                    stock['name'],
                    stock['market_cap'] / 1e9
                ])
                # Color removed rows red
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Format columns
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        
        # Format numbers
        for row in ws.iter_rows(min_row=2):
            if row[0].value and isinstance(row[0].value, date):
                row[0].number_format = 'YYYY-MM-DD'
            if row[4].value and isinstance(row[4].value, (int, float)):
                row[4].number_format = '#,##0.00'
    
    def _add_summary_sheet(self, wb: Workbook, start_date: date, end_date: date):
        """Add summary sheet as the first sheet."""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = "Equal-Weighted Stock Index Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Date range
        ws.merge_cells('A3:E3')
        ws['A3'] = f"Period: {start_date} to {end_date}"
        ws['A3'].font = Font(size=12)
        ws['A3'].alignment = Alignment(horizontal="center")
        
        # Get summary data
        try:
            perf_data = index_service.get_performance(start_date, end_date)
            changes_data = index_service.get_composition_changes(start_date, end_date)
            
            # Performance summary
            ws['A5'] = "Performance Summary"
            ws['A5'].font = Font(bold=True, size=14)
            
            summary = perf_data['summary']
            ws['A6'] = "Total Return:"
            ws['B6'] = f"{summary['total_return']:.2f}%"
            ws['A7'] = "Sharpe Ratio:"
            ws['B7'] = f"{summary['sharpe_ratio']:.2f}"
            ws['A8'] = "Volatility (Daily):"
            ws['B8'] = f"{summary['volatility']:.2f}%"
            ws['A9'] = "Trading Days:"
            ws['B9'] = perf_data['total_days']
            
            # Composition changes summary
            ws['A11'] = "Composition Changes Summary"
            ws['A11'].font = Font(bold=True, size=14)
            
            ws['A12'] = "Total Change Dates:"
            ws['B12'] = changes_data['total_change_dates']
            
            total_additions = sum(len(c['stocks_added']) for c in changes_data['changes'])
            total_removals = sum(len(c['stocks_removed']) for c in changes_data['changes'])
            
            ws['A13'] = "Total Stocks Added:"
            ws['B13'] = total_additions
            ws['A14'] = "Total Stocks Removed:"
            ws['B14'] = total_removals
            
        except ValueError as e:
            ws['A5'] = f"Error loading data: {str(e)}"
        
        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        
        # Add borders to summary sections
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=5, max_row=14, min_col=1, max_col=2):
            for cell in row:
                if cell.value:
                    cell.border = thin_border


# Global service instance
export_service = ExportService()